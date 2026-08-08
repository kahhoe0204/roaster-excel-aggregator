from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill

_PALETTE = [
    "FFF2CC", "D9EAD3", "CFE2F3", "F4CCCC",
    "EAD1DC", "D0E0E3", "FCE5CD", "D9D2E9",
]


def _month_tab(date_str):
    try:
        return datetime.strptime(f"{date_str} 2000", "%d-%b %Y").strftime("%B")
    except ValueError:
        return "Other"


def _branch(source):
    return source.split(" / ", 1)[0]


def _autosize_columns(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value or "")))
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width + 10


def rows_to_xlsx(rows):
    wb = Workbook()
    wb.remove(wb.active)

    branch_colors = {}

    def fill_for(branch):
        color = branch_colors.setdefault(branch, _PALETTE[len(branch_colors) % len(_PALETTE)])
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    sheets = {}
    for row in rows:
        month = _month_tab(row["date"])
        ws = sheets.get(month)
        if ws is None:
            ws = wb.create_sheet(title=month)
            ws.append(["Name", "Date", "Day", "Time", "Source", "Operation Period", "Remark"])
            sheets[month] = ws
        ws.append([
            row["name"].upper(), row["date"], row.get("day"), row["hours"],
            row["source"], row.get("operation_hours"), row.get("remark"),
        ])
        for cell in ws[ws.max_row]:
            cell.fill = fill_for(_branch(row["source"]))

    if not sheets:
        wb.create_sheet(title="Report").append(
            ["Name", "Date", "Day", "Time", "Source", "Operation Period", "Remark"]
        )

    for ws in wb.worksheets:
        _autosize_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
