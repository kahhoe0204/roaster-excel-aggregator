from io import BytesIO
from openpyxl import load_workbook
from app import excel_export


def test_rows_to_xlsx_roundtrip():
    rows = [
        {"name": "Alice", "date": "1-Aug", "day": "Friday", "hours": 9.5, "source": "Branch A / August", "operation_hours": "10:00 AM - 10:00 PM", "remark": "bank in"},
        {"name": "Alice", "date": "3-Aug", "day": "Sunday", "hours": 12.0, "source": "Branch A / August", "operation_hours": "10:00 AM - 10:00 PM", "remark": ""},
    ]
    data = excel_export.rows_to_xlsx(rows)

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["August"]
    ws = wb.active
    values = [[cell.value for cell in row] for row in ws.iter_rows()]

    assert values[0] == ["Name", "Date", "Day", "Time", "Source", "Operation Period", "Remark"]
    assert values[1] == ["ALICE", "1-Aug", "Friday", 9.5, "Branch A / August", "10:00 AM - 10:00 PM", "bank in"]
    assert values[2] == ["ALICE", "3-Aug", "Sunday", 12.0, "Branch A / August", "10:00 AM - 10:00 PM", None]


def test_rows_to_xlsx_empty_rows_still_has_header():
    data = excel_export.rows_to_xlsx([])
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    values = [[cell.value for cell in row] for row in ws.iter_rows()]
    assert values == [["Name", "Date", "Day", "Time", "Source", "Operation Period", "Remark"]]


def test_rows_to_xlsx_splits_months_into_separate_tabs():
    rows = [
        {"name": "Alice", "date": "20-Jul", "day": "", "hours": 9.0, "source": "Branch A / July", "operation_hours": None},
        {"name": "Alice", "date": "1-Aug", "day": "", "hours": 9.5, "source": "Branch A / August", "operation_hours": None},
    ]
    data = excel_export.rows_to_xlsx(rows)
    wb = load_workbook(BytesIO(data))

    assert wb.sheetnames == ["July", "August"]
    assert [c.value for c in wb["July"][2]] == ["ALICE", "20-Jul", None, 9.0, "Branch A / July", None, None]
    assert [c.value for c in wb["August"][2]] == ["ALICE", "1-Aug", None, 9.5, "Branch A / August", None, None]


def test_rows_to_xlsx_autosizes_columns_to_max_value_length():
    rows = [
        {"name": "Alice", "date": "1-Aug", "day": "", "hours": 9.0, "source": "A Very Long Branch Name / August", "operation_hours": None},
    ]
    data = excel_export.rows_to_xlsx(rows)
    wb = load_workbook(BytesIO(data))
    ws = wb["August"]

    assert ws.column_dimensions["E"].width == max(len("Source"), len("A Very Long Branch Name / August")) + 10


def test_rows_to_xlsx_colors_rows_by_branch():
    rows = [
        {"name": "Alice", "date": "1-Aug", "day": "", "hours": 9.0, "source": "Branch A / August"},
        {"name": "Bob", "date": "1-Aug", "day": "", "hours": 9.0, "source": "Branch B / August"},
        {"name": "Alice", "date": "2-Aug", "day": "", "hours": 9.0, "source": "Branch A / August"},
    ]
    data = excel_export.rows_to_xlsx(rows)
    wb = load_workbook(BytesIO(data))
    ws = wb["August"]

    row_a1 = ws[2][0].fill.start_color.rgb
    row_b = ws[3][0].fill.start_color.rgb
    row_a2 = ws[4][0].fill.start_color.rgb

    assert row_a1 == row_a2
    assert row_a1 != row_b
